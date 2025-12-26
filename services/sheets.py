import asyncio
import logging
import os
import openpyxl
from openpyxl import Workbook
import yadisk
from yadisk.exceptions import LockedError
from config import YANDEX_TOKEN, EXCEL_FILE, YANDEX_DIR, REMOTE_PATH_SUBS

logger = logging.getLogger(__name__)
file_lock = asyncio.Lock()

try:
    y = yadisk.YaDisk(token=YANDEX_TOKEN)
except Exception as e:
    y = None

class CloudUploadError(Exception):
    pass

def _ensure_remote_dir_exists(client: yadisk.YaDisk, path: str):
    parts = path.strip("/").split("/")
    current_path = ""
    for part in parts:
        current_path += f"/{part}"
        try:
            if not client.exists(current_path): client.mkdir(current_path)
        except Exception: pass

def _set_column_widths(ws):
    widths = {'A': 18, 'B': 15, 'C': 20, 'D': 20, 'E': 35, 'F': 50, 'G': 18, 'H': 25, 'I': 12}
    for col_letter, width in widths.items():
        try: ws.column_dimensions[col_letter].width = width
        except Exception: pass

def _update_headers_if_needed(ws, new_headers: list):
    try:
        current_headers = [cell.value for cell in ws[1]]
        if current_headers != new_headers:
            for col_num, header in enumerate(new_headers, 1):
                ws.cell(row=1, column=col_num, value=header)
            _set_column_widths(ws)
    except Exception: pass

def _save_to_excel_sync(filename: str, remote_path: str, data: list, headers: list):
    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        _set_column_widths(ws)
        wb.save(filename)
    try:
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        _update_headers_if_needed(ws, headers)
        ws.append(data)
        wb.save(filename)
    except PermissionError:
        raise IOError(f"Файл {filename} открыт.")

    if not y: return
    try:
        if not y.check_token(): raise CloudUploadError("Invalid Token")
        _ensure_remote_dir_exists(y, YANDEX_DIR)
        try:
            y.upload(filename, remote_path, overwrite=True)
        except LockedError:
            y.remove(remote_path)
            import time; time.sleep(1)
            y.upload(filename, remote_path, overwrite=True)
    except Exception as e:
        if isinstance(e, CloudUploadError): raise e
        raise CloudUploadError(f"Upload fail: {e}")

async def add_subscription(user_data: list):
    headers = ["Дата", "User ID", "Username", "Тип подписки", "ФИО", "Способ получения / Доставка", "Телефон", "Выбранные номера", "Согласие ПД"]
    async with file_lock:
        await asyncio.to_thread(_save_to_excel_sync, EXCEL_FILE, REMOTE_PATH_SUBS, user_data, headers)

def find_last_subscription(user_id: int):
    """Ищет запись СТРОГО по новой структуре колонок."""
    if not os.path.exists(EXCEL_FILE):
        return None
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)
        ws = wb.active
        target_id = str(user_id).strip()
        found_data = None
        for row in reversed(list(ws.iter_rows(min_row=2, values_only=True))):
            if not row or len(row) < 5 or row[1] is None:
                continue
            row_id = str(row[1]).strip()
            if row_id.endswith(".0"): row_id = row_id[:-2]
            if row_id == target_id:
                name = row[4]
                phone = str(row[6]) if len(row) > 6 and row[6] else "Не указан"
                delivery_full = str(row[5]) if len(row) > 5 and row[5] else ""
                address = ""
                if "Адрес:" in delivery_full:
                    parts = delivery_full.split("Адрес:")
                    if len(parts) > 1: address = parts[1].strip()
                if name:
                    found_data = {"name": str(name), "phone": phone, "address": address}
                    break
        wb.close()
        return found_data
    except Exception as e:
        logger.error(f"Ошибка чтения истории: {e}")
        return None